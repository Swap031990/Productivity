from py_jama_rest_client.client import JamaClient
import pandas as pd
import numpy as py
import openpyxl

# Setup your Jama instance url, username, and password.
# You may use environment variables, or enter your information directly.
# Reminder: Follow your companies security policies for storing passwords.

# put - edits exitings items
# post - creates new items,relationships etc
# get - gives info on exisiting items

#################################################################################################################################
#   
## Working code to Link item as downlink
#
# Downstream_work function - A. Does Downstream Linking From "From_Item" to "To_Item"
#                            B. Create Excel with Customer API ID to Downlink Req API ID two column required
#                            C. Excel Column will be read and Empty Cell will be Dropped , Dropped ID will be printed
#                            D. 
#
#
#################################################################################################################################


# Create the JamaClient
jama_client = JamaClient('https://luminar.jamacloud.com', credentials=('kwxt2qcnup2eyky', 'w8knzuhhoqqmc1h1dc34gnnzn'), oauth=True)

# Get the list of projects from Jama
# The client will return to us a JSON array of Projects, where each project is a JSON object.
project_list = jama_client.get_projects()

#

# Print the data out for each project.
def proj():
    for project in project_list:
        project_name = project['fields']['name']
        print('\n---------------' + project_name + '---------------')

        # Print each field
        for field_name, field_data in project.items():

            # If one of the fields(i.e. "fields") is a dictionary then print it's sub fields indented.
            if isinstance(field_data, dict):
                print(field_name + ':')
                # Print each sub field
                for sub_field_name in field_data:
                    sub_field_data = field_data[sub_field_name]
                    print('\t' + sub_field_name + ': ' + str(sub_field_data))

            # If this field is not a dictionary just print its field.
            else:
                    print(field_name + ': ' + str(field_data))

#################################################################################################################################
# dropna - will drop NAN - not a number or empty fields in excel for given column
def Downlink():
    #Reading the excel sheet skipping unwanted rows..
    requirements = pd.read_excel("C:\Swapnil\Tool_To_Install\py-jama-rest-client-master\py-jama-rest-client-master\examples\MB_Supplemental_Master.xlsx",skiprows=3)
    #Skipping blank values if any
    requirements.dropna(0,subset=['API ID','MB_API_ID'],inplace=True)
    for index,row in requirements.iterrows():
            # Handling Error if link already exists
            try:
                    print(int(row['MB_API_ID']),int(row['API ID']))
                    #Linking API with Relationtype as per json -> Related to here
                    s = jama_client.post_relationship(int(row['MB_API_ID']),int(row['API ID']),4)
                    print(s)
            except:
                    print(int(row['MB_API_ID']),"---->",int(row['API ID']),"Link Already Exists")


#################################################################################################################################
### DO NOT CHANGE BELOW CODE  ONLY FOLLOWING CHANGES ARE APPLICABLE######
# CHANGE FOR LOOP RANGE
# CHANGE FROM_ITEM AND TO_ITEM COLUMN NO
def Downstream_work():
    path = "C:\Swapnil\Tool_To_Install\py-jama-rest-client-master\py-jama-rest-client-master\examples\LinkExcel.xlsx"
    exc_obj = openpyxl.load_workbook(path)

    sheet_obj = exc_obj.active

    row = sheet_obj.max_row
    column = sheet_obj.max_column
    
    print("Total Rows:", row)
    print("Total Columns:", column)
    
"""
    # Cell object is created by using 
    # sheet object's cell() method. 
    for i in range(7, row + 1):     #RANGE CAN BE CHANGED
        cell_obj = sheet_obj.cell(row = i, column = 2)   # TO item column number MAY CHANGE
        To_item = cell_obj.value
        cell_obj = sheet_obj.cell(row = i, column = 6)   # From Item Column MAY CHANGE
        From_item = cell_obj.value

        # Print value of cell object 
        # using the value attribute 
        if From_item is None:
            print("Dropped",To_item,From_item)
        else:
            print("From_item = ",From_item,end=' ') 
            print("To_item = ",To_item,end=' ')
            try:
                #s = jama_client.post_relationship(From_item,To_item,4)  # COMMENT THIS LINE TO UNDERSTAND IF CORRECT COLUMN NO SELECTED
                #print(s)
                print("Linked") 
            except:
                print("Link Already Exists")
"""
  
###### item type ######
# 32 - sub folder
# 33 - text item
#  149 - sensor requirements

# MB supp
#item_id_jama = 1434196
#project_id_jama = 333

#System Sandbox
item_id_jama = 2255763
project_id_jama = 406

# 1  Able to read items and fields
items = jama_client.get_item(item_id_jama)
#print(items)
#print(items["globalId"])
#print(items["fields"]["name"])


#proj()

# 2 LOCK STATE of item
lock_items = jama_client.get_item_lock(1984494)
#print(lock_items)

# 3 LOCK STATE Update item and UNLOCK State of an item
#locked_items = jama_client.put_item_lock(1984494,'True')
#print(locked_items)

#edit fileds  , doesnt work if we have multiple fields with same name fetches only first one
#items["fields"]["description"] = 'Sensor Requirements for Communication Subsystem'
#jama_client.put_item(406,1984494,33,child_item_type_id=0,location={'item': 1984492},fields=items["fields"])

#locked_items = jama_client.put_item_lock(1984494,'False')
#print(locked_items)


# 4 get Upstream and Downstream of a item

#upstream  - fromitem:upstream id, toItem:Current item, upstream type = 4 , type: 'relationships'
Uprelationships = jama_client.get_items_upstream_relationships(item_id_jama,1)
#print(Uprelationships)

#Downstream
Downrelationships = jama_client.get_items_downstream_related(item_id_jama,1)
#print(Downrelationships)

 #### put upstream to ID
#Uprela = jama_client.post_relationship(1990811,1984494,4)
#Uprela = jama_client.post_relationship(373720,2015013,4)
#Downlink()   #first script basic to link downlink
#Uprela = jama_client.get_items_upstream_relationships(item_id_jama,1)
Uprela = jama_client.get_items_downstream_relationships(item_id_jama,1)
print(Uprela)

#when multiple downstream links are there from different project can be filtered with toItem and documentkey
#print(Uprela[0]['toItem'])
#print(Uprela[1]['toItem'])
#print(Uprela[2]['toItem'])

'''items0 = jama_client.get_item(Uprela[0]['toItem'])
print(items0["documentKey"])
items2 = jama_client.get_item(Uprela[2]['toItem'])
print(items2["documentKey"])'''


for i in range(3):
     print(Uprela[i]['toItem'])
     items0 = jama_client.get_item(Uprela[i]['toItem'])
     

     if 'IS1-LS_RQ-986' == items0["documentKey"]:
          print(items0["documentKey"])
          DelRela_items = jama_client.delete_relationships(Uprela[i]['id'])  
          print(DelRela_items)

#uncomment code inside it
#Downstream_work()


# 5 Get tags in project
Tag_list = jama_client.get_tags(project_id_jama,1)
#print(Tag_list)

# 6 tagged items  tag id present in jama weblink
#Tag_items = jama_client.get_tagged_items(tag_id=592)
#Tag_items = jama_client.get_tagged_items(tag_id=637)
#print(Tag_items)

#6.1 Add existing tag to new requirements 
#TagNew_items = jama_client.post_item_tag(item_id_jama,tag_id=637)
#print(TagNew_items)

# 7  Get all relationship items id
Rel_items = jama_client.get_relationship_types()
#print(Rel_items)

#8 get relationships
# relationshipType = 6  Downlink  = 21 ??
# relationtype :  4 = related to
# 21 = Mapping

Rela_items = jama_client.get_relationship(relationship_id=333)
#print(Rela_items)


#9 Delete relationships need relationship id
# get relationship id using jama_client.get_items_downstream_relationships(item_id_jama,1) for upstream also
#DelRela_items = jama_client.delete_relationships(410714)  
#print(DelRela_items)


#10 Edit existing rela
#PutRela_items = jama_client.put_relationship(4,item_id_jama,item_id_jama+1) # not working
#print(PutRela_items)


#11 
#GetRela_items = jama_client.get_relationship_type(382976) # not working
#print(GetRela_items)


#print(items)