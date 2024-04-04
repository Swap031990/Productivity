# MGui.py
# Revision : 1.8.2
# Author   : Swapnil S. Bhujbal
# Email ID : Swapnil.bhujbal@luminartech.com
# Date     : 10/02/2024
# Details  : Difference between gui.py and this file is here two separate window layouts are used; This file is developed on base of 1.5 
#            version of gui.py
#
# Revision History : 1.0 - Able to Link to Jama items with Downstream relationship
#                    1.1 - Menu Bar added
#                    1.2 - Linking via excel column one column From item API_ID to To item API_ID
#                    1.3 - Radio option added for selecting of Uplink or Downlink linking
#                    1.4 - Excel browse path added to link two column of API_IDs
#                    1.5 - Aesthetic changed and Demo linking via excel option added
#                    1.6 - Option to add users jama credentials
#                    1.6.1 - Issue about jama credentials fetching from settings resolved
#                    1.7   - Verification criteria of a single item can be updated , added check if verification criteria is blank then add key to dictionary s
#                    1.8   - same Verification criteria updation to multiple item
#                    1.8.1 - 1:1 verification criteria , unique vc as per excel column, column 2 - API ID and Column 3 verification scriteria
#                    1.8.2 - resolved issue in excel verification criteria updation from previous release
###################################################################################################################
import pandas as pdprint
import numpy as py
import openpyxl
import PySimpleGUI as sg
import time
import json
from py_jama_rest_client.client import JamaClient


#Debugger Console
Debug = 'disable'

#Dummy credentials
client_ID = 'ABCDEFGH'
client_code = 123456789

#jama_client = 0
#################################################################################################################################
### DO NOT CHANGE BELOW CODE  ONLY FOLLOWING CHANGES ARE APPLICABLE######
# CHANGE FOR LOOP RANGE
# CHANGE FROM_ITEM AND TO_ITEM COLUMN NO
def Downstream_work(Expath,from_id_col,To_id_Col):
    global window
    global values
    #path = "C:\Swapnil\Tool_To_Install\py-jama-rest-client-master\py-jama-rest-client-master\examples\LinkExcel.xlsx"

   # exc_obj = openpyxl.load_workbook(path)
    exc_obj = openpyxl.load_workbook(Expath)
    if Debug == 'enable':
        sg.Print(Expath)
        sg.Print(from_id_col)
        sg.Print(To_id_Col)
    sheet_obj = exc_obj.active

    row = sheet_obj.max_row
    column = sheet_obj.max_column
    if Debug == 'enable':
        sg.Print("Total Rows:", row)
        sg.Print("Total Columns:", column)


    ####### progress bar update ############
    """
    for i in range(100):
        window['-PBAR-'].update(current_count=i + 1)
        window['-OUT-'].update(str(i + 1))
        #time.sleep(1)
    """


    # Cell object is created by using 
    # sheet object's cell() method. 
    for i in range(3, row + 1):     #RANGE CAN BE CHANGED
        cell_obj = sheet_obj.cell(row = i, column = int(To_id_Col))   # TO item column number MAY CHANGE
        To_item = cell_obj.value
        cell_obj = sheet_obj.cell(row = i, column = int(from_id_col))   # From Item Column MAY CHANGE
        From_item = cell_obj.value

    ####### progress bar update ############
        window['-PBAR-'].update(current_count=(((i + 1)/row)*100))
        window['-OUT-'].update(str(str((int)((i/row)*100))+"%"))
        #time.sleep(1)

        # Print value of cell object 
        # using the value attribute 
        if From_item is None:
            if Debug == 'enable':
                sg.Print("Dropped",To_item,From_item)
        else:
            if Debug == 'enable':
                sg.Print("From_item = ",From_item,end=' ') 
                sg.Print("To_item = ",To_item,end=' ')
            try:
                if values['Active'] == True:
                    #sg.Print("Properly") 
                    s = jama_client.post_relationship(From_item,To_item,4)  # COMMENT THIS LINE TO UNDERSTAND IF CORRECT COLUMN NO SELECTED
                if Debug == 'enable':
                    #sg.Print(s)
                    sg.Print("Linked") 
            except:
                if Debug == 'enable':
                    sg.Print("Either Link Already Exists or check Jama credentials")
    sg.Print("Linking Finished")
    Finished = 1
    end = time.time()
    elapsed = end-start
    sg.Print("Elapsed time for ",row,"requirement linking is =",elapsed)
    ############ reset progress bar ############
    time.sleep(3)
    window['-PBAR-'].update(current_count=0)
    window['-OUT-'].update(str(str(0)+"%"))
  
################## Downlink individual item ###########################
def Downlink(from_item,to_item):
   #sg.popup_error("In Downlink")
    if Debug == 'enable':
        sg.Print("From_item = ",from_item,end=' ') 
        sg.Print("To_item = ",to_item)
    try:
        s = jama_client.post_relationship(from_item,to_item,4)  # GUI crashes here because of entered incorrect credentials
        #sg.popup(title="Successful",custom_text="!! Linking Done !!")
        sg.popup_no_buttons('Linking Done',title='Successful')
        if Debug == 'enable':
            sg.Print(s)
            sg.Print("Linked") 
    except:
        sg.popup(title="Error",custom_text="!! Downlink already exists !!")
        if Debug == 'enable':
            sg.Print("Either Link Already Exists or check Jama credentials")
################## Verification criteria update - individual item ###########################
def Update_VC(Api_item,Ver_Cri):

    global client_ID
    global client_code
    jama_client2 = JamaClient('https://luminar.jamacloud.com', credentials=(client_ID, client_code), oauth=True)
    sg.Print("Api_item = ",Api_item) 
    sg.Print("Ver_Cri = ",Ver_Cri)

    s = jama_client2.get_item(Api_item)
    criteria_new = s

    # To print ID from S dictinary
    #sg.Print(s['id'])  
    #Prints OLD Verification criteria
    sg.Print("Old Verification criteria",end=' ') # To print two statements in same line  
    if 'verification_critera$149'  in s['fields']:
        #sg.Print("Present Verification criteria") 
        pass
    else:
        s['fields']['verification_critera$149'] = "NA"

    sg.Print(s['fields']['verification_critera$149'])  
    #sg.Print(s['fields']['documentKey'])  
    
    #arguments to pass to put function
    """sg.Print(s['location']['parent']['item']) 
    sg.Print(s['project']) 
    sg.Print(s['itemType']) """

    parent_location_ID = s['location']['parent']['item']
    project_ID = s['project']
    ItemTyope_ID = s['itemType']

    #Prints New Verification criteria update in variable
    #check if key ( dictionary has key and value like a table) verification criteria is not available
    if "['fields']['verification_critera$149']" not in criteria_new:
        criteria_new['fields']['verification_critera$149'] = "NA"
    else:
        sg.Print("Present Verification criteria")  

    criteria_new['fields']['verification_critera$149'] = Ver_Cri
    #sg.Print("New Verification criteria",end=' ')  
    sg.Print(criteria_new['fields']['verification_critera$149'])  

    # each item need item type,project id,parent api id
    #jama_client2.put_item(406,Api_item,149,None,2270911,criteria_new['fields'])
    jama_client2.put_item(project_ID,Api_item,ItemTyope_ID,None,parent_location_ID,criteria_new['fields'])

    """try:
        s = jama_client2.put_item(406,Api_item,149,None,2270911,criteria_new['fields'])  # GUI crashes here because of entered incorrect credentials
        sg.popup_no_buttons('Linking Done',title='Successful')
        sg.Print(s)
 
    except:
        sg.popup(title="Error",custom_text="!! Failed to Update Verification criteria !!")
        if Debug == 'enable':
            sg.Print("Either Link Already Exists or check Jama credentials")"""
################################################################################################################################



################## Verification criteria update - individual item ###########################
def Update_VC_Multiple(fpath,Api_item,Ver_Cri):

    global client_ID
    global client_code

    # exc_obj = openpyxl.load_workbook(path)
    excel_obj = openpyxl.load_workbook(fpath)
    sheet_obj = excel_obj.active
    row = sheet_obj.max_row
    #column = sheet_obj.max_column
    if Ver_Cri == "":
        return 0

    Final_VC = Ver_Cri

    # Cell object is created by using 
    # sheet object's cell() method. 
    for i in range(4, row + 1):     #RANGE CAN BE CHANGED

        #read 2 column
        cell_obj = sheet_obj.cell(row = i, column = int(2))   # TO item column number MAY CHANGE
        Api_item = cell_obj.value


        #fetch 3 rd column
        if Ver_Cri.upper() == 'NA':
            cell_obj1 = sheet_obj.cell(row = i, column = int(3))   # TO item column number MAY CHANGE
            Final_VC = cell_obj1.value
            #sg.Print("Api_item = ",Api_item,end=' ') # To print two statements in same line
            sg.Print("Excel Ver_Cri = ",Final_VC)

        if Api_item == None:
            break
        else:  
            Update_VC(Api_item,Final_VC)


    sg.Print(" Number of items updated =",i-4)
    return 1
################################################################################################################################
    
################## Delete a relationship item ###########################
# to delete a relationship , you need to get relatinship id which is different than API ID , using GET fetch all relationship ID then based on Fromitem and Toitem filter and delete with 
# relationship ID
#  GET(project_ID)  - result of all relationship in project       "id": 382976,
# #     "fromItem": 2076999,
#  #    "toItem": 2077005,
#  Delete(382976)
def Delete_Relationship():
    jama_client1 = JamaClient('https://luminar.jamacloud.com', credentials=(client_ID, client_code), oauth=True)
    try:
        #s = jama_client1.delete_relationships(1984494)
        #s = jama_client1.get_relationship(382976)
        #s = jama_client1.get_items_downstream_related(2076999,1)  # this returns directly API ID of Downstream link item and every info about it

        #s = jama_client1.get_items_downstream_relationships(2076999,1) #this returns downstream relationship ID using API ID of an item with following infor
        #[{'id': 382976, 'fromItem': 2076999, 'toItem': 2077005, 'relationshipType': 21, 'suspect': False, 'type': 'relationships'}]
        s = jama_client1.get_items_upstream_relationships(2077005,1) #this returns upstream relationship ID using API ID of an item with following infor
        #{'id': 382976, 'fromItem': 2076999, 'toItem': 2077005, 'relationshipType': 21, 'suspect': False, 'type': 'relationships'},
        sg.Print(s)

    except:
        sg.popup(title="Error",custom_text="!! Delete failure !!")
#################################################################################################################################
#################################################################################################################################
#################################################################################################################################
#################################################################################################################################
#################################################################################################################################
#################################################################################################################################


############ issue after launching .exe in others comp ###################
def win2():
    global client_ID
    global client_code
    global client_Name
    global settings
    #global jama_client
    sg.set_options(font=('Arial Bold', 16))
    layout = [
    [sg.Text('Settings', justification='left')],
    [sg.Text('User name', size=(10, 1), expand_x=True),
    sg.Input(key='-USER-')],
    [sg.Text('Client_ID', size=(10, 1), expand_x=True),
    sg.Input(key='-Client_ID-')],
    [sg.Text('Client_Code', size=(10, 1), expand_x=True),
    sg.Input(key='-Client_Code-')],
    [sg.Button("LOAD"), sg.Button('SAVE'), sg.Button('Exit')]
    ]
    window1 = sg.Window('Jama Credentials', layout, size=(715, 200))
    # Event Loop
    while True:
        event, values = window1.read()

        
        if event in (sg.WIN_CLOSED, 'Exit'):
            break


        if event == 'LOAD':
            f = open("settings.txt", 'r')
            settings = json.load(f)
            window1['-USER-'].update(value=settings['-USER-'])
            window1['-Client_ID-'].update(value=settings['-Client_ID-'])
            window1['-Client_Code-'].update(value=settings['-Client_Code-'])

            client_Name = settings['-USER-']
            client_ID = settings['-Client_ID-']
            client_code = settings['-Client_Code-']

            """sg.Print("User Name=",settings['-USER-'])
            sg.Print("client ID=",settings['-Client_ID-'])
            sg.Print("client Secret code=",settings['-Client_Code-'])"""


        if event == 'SAVE':
            settings = {'-USER-': values['-USER-'],
            '-Client_ID-': values['-Client_ID-'],
            '-Client_Code-': values['-Client_Code-']}
            f = open("settings.txt", 'w')
            json.dump(settings, f)
            f.close()
            
            client_Name = settings['-USER-']
            client_ID = settings['-Client_ID-']
            client_code = settings['-Client_Code-']

            """"sg.Print("User Name=",settings['-USER-'])
            sg.Print("client ID=",settings['-Client_ID-'])
            sg.Print("client Secret code=",settings['-Client_Code-'])"""

    window1.close()


###################################################################################################################
def win3():
    global API_ID
    global Verification_criteria

    #global jama_client
    sg.set_options(font=('Arial Bold', 16))
    layout = [
    [sg.Text('Verification Criteria single or Multiple', justification='left')],
    [sg.Radio("Single","ena",key='Single',default=True),sg.Radio("Multiple","ena",key='Multiple')],
    [sg.Text("Browse excel with API_ID", size=(19,1),tooltip="Column 2:API ID & Column 3: Verification criteria"),sg.Input(key="-PATH-"),sg.FilesBrowse()],
    [sg.Text('API ID', size=(10, 1), expand_x=True),
    sg.Input(key='-API_ID-')],
    [sg.Text('Verification criteria', size=(10, 1), expand_x=True,tooltip="Put NA if excel has verification criteria"),
    sg.Input(key='-Verification_criteria-')],
    [sg.Button("Update"), sg.Button('Exit')]
    ]

    # Window size (x, y)  x = 950 horizontal and y = 250 vertical
    window2 = sg.Window('Verification criteria', layout, size=(950, 250))



    # Event Loop
    while True:
        event, Data = window2.read()

        #sg.Print("API ID=", Data['-API_ID-'])
        #sg.Print("Verification criteria=", Data['-Verification_criteria-'])    

        if event in (sg.WIN_CLOSED, 'Exit'):
            break

        if event == 'XXX':
            # read API ID Jama item and update verification criteria
            sg.Print("Reserved for Future Release")

        if event == 'Update':

            if Data['Single'] == True:
                # read API ID Jama item and update verification criteria
                sg.Print("Updation of A Verification criteria Started")
                Update_VC(Data['-API_ID-'],Data['-Verification_criteria-'])
            else:
                # read API ID Jama item and update verification criteria
                sg.Print("Updation of Multiple Verification criteria Started")
                success = Update_VC_Multiple(Data['-PATH-'],Data['-API_ID-'],Data['-Verification_criteria-'])
                sg.Print("Operation successfull =",success)


    window2.close()


###################################################################################################################
### fetch default setting values ###
f = open("settings.txt", 'r')
settings = json.load(f)

API_ID = 0
Verification_criteria = "NA"
client_Name = settings['-USER-']
client_ID = settings['-Client_ID-']
client_code = settings['-Client_Code-']

menu_def = [['File', ['New', 'Revision History', 'Contact', 'Exit', ]], ['Setting', ['Fetch_Jama_Cred','Debug Window', 'options', 'jama_credentials','update verification criteria'], ],  ['Help', 'About...'], ]
sg.theme('DarkTeal4')

layout = [[sg.Menu(menu_def)],
    [sg.Image('unnamed.png',expand_x=True, expand_y=True)],
    [sg.Text("Deactivate jama linking", size=(19,1)),sg.Radio("Yes","ena",key='Deactive',default=True),sg.Radio("No","ena",key='Active'),sg.Text(" In case excel by-default its Downlink, Reverse col no to uplink", text_color='Orange',font=('Arial Bold', 15),expand_x=True)],
    [sg.Text("Browse excel with API_ID", size=(19,1)),sg.Input(key="-IN-"),sg.FilesBrowse()],
    [sg.Text("API_ID From Col no", size=(19,1)),sg.Input(key="-IN-3")],
    [sg.Text("API_ID To Col no", size=(19,1)),sg.Input(key="-IN-4"),sg.Button(" START ",tooltip='Start Linking')],
    #[sg.Text("Output Folder:"),sg.Input(key="-OUT-"),sg.FolderBrowse()],
    #[sg.Text("Get Downstream Relationship ID:"),sg.Input(key="-IN-3"),sg.Button(" Get "),sg.Button(" Put ")],
    #[sg.Text("Downstream linking")],
    [sg.Text("Linking Type", size=(19,1)),sg.Radio("Downstream","gen",key='Downstream',default=True),sg.Radio("Upstream","gen",key='Upstream')],
    [sg.Text("API_ID From:",size=(19,1)),sg.Input(key="-IN-1")],
    [sg.Text("API_ID To:",size=(19,1)),sg.Input(key="-IN-2"),sg.Button(" Link ")],
    [sg.ProgressBar(100, orientation='h', bar_color=("green","grey"),expand_x=True, size=(20, 20),  key='-PBAR-')],
    [sg.Text('', key='-OUT-', enable_events=True, font=('Arial Bold', 16), justification='center', expand_x=True)],
    [sg.Text("",size=(53,1)),sg.Button("CLOSE")]

]

margins=(100, 50)
#create a window
window = sg.Window("Luminar Jama Requirement Link",layout,margins)


# Create the JamaClient
#jama_client = JamaClient('https://luminar.jamacloud.com', credentials=('kwxt2qcnup2eyky', 'w8knzuhhoqqmc1h1dc34gnnzn'), oauth=True)
#jama_client = JamaClient('https://luminar.jamacloud.com', credentials=(client_ID, client_code), oauth=True)

#create a loop
while True:
    event, values = window.read()

    if event == sg.WIN_CLOSED:
        break
    if event == "API_ID From":  # can we do event for text , no we cant 
        sg.popup_error("Please enter correct ID")
    if event == "CLOSE":
        window.close()

#################################### events  #############################################################################        
    if event == " Get ":
        if Debug == 'enable':
            sg.Print(values["-IN-3"])
        #Downstream
        Downrelationships = jama_client.get_items_downstream_related(values["-IN-3"],1)
        if Debug == 'enable':
            sg.Print(Downrelationships)
    if event == " Put ":
        if Debug == 'enable':
            sg.Print(values["-IN-3"])
        #Downstream
        Prelationships = jama_client.put_relationship(2015026,values["-IN-3"],values["-IN-2"],4)
        if Debug == 'enable':
            sg.Print(Prelationships)

    if event == " Link ":
        jama_client = JamaClient('https://luminar.jamacloud.com', credentials=(client_ID, client_code), oauth=True)
        if Debug == 'enable':
            sg.Print("link started")

        #Basically only downlink function is enough for both uplink and downlink, for uplink just reverse jama item api ids
        if values['Upstream'] == True:
            if Debug == 'enable':
                sg.Print("Upstream selected")
            Current_item=values["-IN-2"]
            Link_item=values["-IN-1"]
        if values['Downstream'] == True:
            if Debug == 'enable':
                sg.Print("Downstream selected")
            Current_item=values["-IN-1"]
            Link_item=values["-IN-2"]

    # Downlink(
    #     from_item=values["-IN-1"],
    #     to_item=values["-IN-2"]
    #     )

        Downlink(Current_item,Link_item)
        
    if event == " START ":
            start = time.time()

            jama_client = JamaClient('https://luminar.jamacloud.com', credentials=(client_ID, client_code), oauth=True)

            if values['Deactive'] == True:
                Debug = 'enable'
                sg.Print("Linking in jama Deactivated")
            Downstream_work(values["-IN-"],values["-IN-3"],values["-IN-4"])

################ Menu Bar events   ################################################################   

#### Future implementations ####
    if event == 'New':
        sg.popup_no_buttons('Reserved for future implementations',title='INFO')
        Delete_Relationship()
    if event == 'options':
#       sg.popup_no_buttons('Reserved for future implementations',title='INFO')
        sg.Print("client ID=",client_ID)
        sg.Print("client Secret code=",client_code)


    if event == 'Fetch_Jama_Cred':
        win2()

        sg.Print("New client Name=",client_Name)
        sg.Print("New client ID=",client_ID)
        sg.Print("New client Secret code=",client_code)

    if event == 'jama_credentials':

        client_Name=sg.popup_get_text("enter a client Name: ")
        client_ID=sg.popup_get_text("enter a client ID: ")
        client_code=sg.popup_get_text("enter a client secret code: ")

        sg.Print("Main client Name=",client_Name)
        sg.Print("Main client ID=",client_ID)
        sg.Print("Main client Secret code=",client_code)

        # No need of following as Link or START event will link IDs there only jamaclient is required
        #try:
            # Update the JamaClient with new credentials
            #jama_client = JamaClient('https://luminar.jamacloud.com', credentials=(client_ID, client_code), oauth=True)
            #data= jama_client.get_items_downstream_related(401046,1)
            #sg.Print(data)
        #except:
            #sg.popup(title="Error",custom_text="!! Please Enter Correct Credentials !!")
    ###############################
    if event == 'update verification criteria':
        win3()

    if event == 'Exit':
        window.close()
    if event == 'Debug Window':
        Debug = 'enable'
        sg.Print('Debugger window started')

    if event == 'About...':
        #print('Current vserion is 1.8.2')
        sg.popup_no_buttons("resolved issue in excel verification criteria updation from previous release",text_color="Yellow",title="Revision 1_8_2",non_blocking=True)

    if event == 'Revision History':
        file=open("Rev_His.txt")
        text=file.read()
        sg.popup_scrolled(text, title="Revision History", font=("Arial Bold", 16), size=(50,10))

    if event == 'Contact':
        sg.popup_no_buttons('swapnil.bhujbal@luminartech.com',title="Conatct me @",text_color="orange", non_blocking=True)
    
window.close()


#####################################################################################################
################################################################################################## 
################################################################################################## 
################################################################################################## 
# this was used to launch first window as main then second window as win2
"""
if __name__ == "__main__":
    main()
"""