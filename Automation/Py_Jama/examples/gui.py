# Gui.py
# Revision : 1.1
# Author   : Swapnil S. Bhujbal
# Email ID : Swapnil.bhujbal@luminartech.com
# Date     : 20/09/2023
#
# Revision History : 1.0 - Able to Link to Jama items with Downstream relationship
#                    1.1 - Menu Bar added
#                    1.2 - Linking via excel column one column From item API_ID to To item API_ID
#                    1.3 - Radio option added for selecting of Uplink or Downlink linking
#                    1.4 - Excel browse path added to link two column of API_IDs
#                    1.5 - Aesthetic changed and Demo linking via excel option added
#                    1.6 - Option to add users jama credentials
###################################################################################################################
import pandas as pdprint
import numpy as py
import openpyxl
import PySimpleGUI as sg
import time
import json
from py_jama_rest_client.client import JamaClient

#################################################################################################################################
### DO NOT CHANGE BELOW CODE  ONLY FOLLOWING CHANGES ARE APPLICABLE######
# CHANGE FOR LOOP RANGE
# CHANGE FROM_ITEM AND TO_ITEM COLUMN NO
def Downstream_work(Expath,from_id_col,To_id_Col):
    #path = "C:\Swapnil\Tool_To_Install\py-jama-rest-client-master\py-jama-rest-client-master\examples\LinkExcel.xlsx"
   # exc_obj = openpyxl.load_workbook(path)
    exc_obj = openpyxl.load_workbook(Expath)
    if Debug == 'enable':
        sg.Print(Expath)
        sg.Print(from_id_col)
        sg.Print(To_id_Col)
    sheet_obj = exc_obj.active

    row = sheet_obj.max_row
    column = sheet_obj.max_column
    if Debug == 'enable':
        sg.Print("Total Rows:", row)
        sg.Print("Total Columns:", column)


    ####### progress bar update ############
    """
    for i in range(100):
        window['-PBAR-'].update(current_count=i + 1)
        window['-OUT-'].update(str(i + 1))
        #time.sleep(1)
    """


    # Cell object is created by using 
    # sheet object's cell() method. 
    for i in range(3, row + 1):     #RANGE CAN BE CHANGED
        cell_obj = sheet_obj.cell(row = i, column = int(To_id_Col))   # TO item column number MAY CHANGE
        To_item = cell_obj.value
        cell_obj = sheet_obj.cell(row = i, column = int(from_id_col))   # From Item Column MAY CHANGE
        From_item = cell_obj.value

    ####### progress bar update ############
        window['-PBAR-'].update(current_count=(((i + 1)/row)*100))
        window['-OUT-'].update(str(str((int)(((i + 1)/row)*100))+"%"))
        #time.sleep(1)

        # Print value of cell object 
        # using the value attribute 
        if From_item is None:
            if Debug == 'enable':
                sg.Print("Dropped",To_item,From_item)
        else:
            if Debug == 'enable':
                sg.Print("From_item = ",From_item,end=' ') 
                sg.Print("To_item = ",To_item,end=' ')
            try:
                if values['Active'] == True:
                    #sg.Print("Properly") 
                    s = jama_client.post_relationship(From_item,To_item,4)  # COMMENT THIS LINE TO UNDERSTAND IF CORRECT COLUMN NO SELECTED
                if Debug == 'enable':
                    #sg.Print(s)
                    sg.Print("Linked") 
            except:
                if Debug == 'enable':
                    sg.Print("Either Link Already Exists or check Jama credentials")
    sg.Print("Finished")
    Finished = 1

    ############ reset progress bar ############
    time.sleep(5)
    window['-PBAR-'].update(current_count=0)
    window['-OUT-'].update(str(str(0)+"%"))
  
################## Downlink individual item ###########################
def Downlink(from_item,to_item):
   #sg.popup_error("In Downlink")
    if Debug == 'enable':
        sg.Print("From_item = ",from_item,end=' ') 
        sg.Print("To_item = ",to_item)
    try:
        s = jama_client.post_relationship(from_item,to_item,4)  # COMMENT THIS LINE TO UNDERSTAND IF CORRECT COLUMN NO SELECTED
        #sg.popup(title="Successful",custom_text="!! Linking Done !!")
        sg.popup_no_buttons('Linking Done',title='Successful')
        if Debug == 'enable':
            sg.Print(s)
            sg.Print("Linked") 
    except:
        sg.popup(title="Error",custom_text="!! Downlink already exists !!")
        if Debug == 'enable':
            sg.Print("Either Link Already Exists or check Jama credentials")

#################################################################################################################################


menu_def = [['File', ['New', 'Revision History', 'Contact', 'Exit', ]], ['Setting', ['fetch','Debug Window', 'options', 'jama_credentials'], ],  ['Help', 'About...'], ]

#Debugger Console
#sg.show_debugger_window(location=(10,10))
Debug = 'disable'

sg.theme('DarkTeal4')

layout = [[sg.Menu(menu_def)],
    [sg.Image('unnamed.png',expand_x=True, expand_y=True)],
    [sg.Text("Deactivate jama linking", size=(19,1)),sg.Radio("Yes","ena",key='Deactive',default=True),sg.Radio("No","ena",key='Active'),sg.Text(" In case excel by-default its Downlink, Reverse col no to uplink", text_color='Orange',font=('Arial Bold', 15),expand_x=True)],
    [sg.Text("Browse excel with API_ID", size=(19,1)),sg.Input(key="-IN-"),sg.FilesBrowse()],
    [sg.Text("API_ID From Col no", size=(19,1)),sg.Input(key="-IN-3")],
    [sg.Text("API_ID To Col no", size=(19,1)),sg.Input(key="-IN-4"),sg.Button(" START ",tooltip='Start Linking')],
    #[sg.Text("Output Folder:"),sg.Input(key="-OUT-"),sg.FolderBrowse()],
    #[sg.Text("Get Downstream Relationship ID:"),sg.Input(key="-IN-3"),sg.Button(" Get "),sg.Button(" Put ")],
    #[sg.Text("Downstream linking")],
    [sg.Text("Linking Type", size=(19,1)),sg.Radio("Downstream","gen",key='Downstream',default=True),sg.Radio("Upstream","gen",key='Upstream')],
    [sg.Text("API_ID From:",size=(19,1)),sg.Input(key="-IN-1")],
    [sg.Text("API_ID To:",size=(19,1)),sg.Input(key="-IN-2"),sg.Button(" Link ")],
    [sg.ProgressBar(100, orientation='h', bar_color=("green","grey"),expand_x=True, size=(20, 20),  key='-PBAR-')],
    [sg.Text('', key='-OUT-', enable_events=True, font=('Arial Bold', 16), justification='center', expand_x=True)],
    [sg.Text("",size=(53,1)),sg.Button("CLOSE")]

]

margins=(100, 50)
#create a window
window = sg.Window("Luminar Jama Requirement Link",layout,margins)

#Dummy credentials
client_ID = 'ABCDEFGH'
client_code = 123456789
# Create the JamaClient
#jama_client = JamaClient('https://luminar.jamacloud.com', credentials=('kwxt2qcnup2eyky', 'w8knzuhhoqqmc1h1dc34gnnzn'), oauth=True)
#jama_client = JamaClient('https://luminar.jamacloud.com', credentials=(client_ID, client_code), oauth=True)

#create a loop
while True:
    event, values = window.read()

    if event == sg.WIN_CLOSED:
        break
    if event == "API_ID From":  # can we do event for text , no we cant 
        sg.popup_error("Please enter correct ID")
    if event == "CLOSE":
        window.close()

#################################### events  #############################################################################        
    if event == " Get ":
        if Debug == 'enable':
            sg.Print(values["-IN-3"])
        #Downstream
        Downrelationships = jama_client.get_items_downstream_related(values["-IN-3"],1)
        if Debug == 'enable':
            sg.Print(Downrelationships)
    if event == " Put ":
        if Debug == 'enable':
            sg.Print(values["-IN-3"])
        #Downstream
        Prelationships = jama_client.put_relationship(2015026,values["-IN-3"],values["-IN-2"],4)
        if Debug == 'enable':
            sg.Print(Prelationships)

    if event == " Link ":
        if Debug == 'enable':
            sg.Print("link started")

        #Basically only downlink function is enough for both uplink and downlink, for uplink just reverse jama item api ids
        if values['Upstream'] == True:
            if Debug == 'enable':
                sg.Print("Upstream selected")
            Current_item=values["-IN-2"]
            Link_item=values["-IN-1"]
        if values['Downstream'] == True:
            if Debug == 'enable':
                sg.Print("Downstream selected")
            Current_item=values["-IN-1"]
            Link_item=values["-IN-2"]

       # Downlink(
       #     from_item=values["-IN-1"],
       #     to_item=values["-IN-2"]
       #     )

        Downlink(
          from_item=Current_item,
          to_item=Link_item
           )
        
    if event == " START ":
            if values['Deactive'] == True:
                Debug = 'enable'
                sg.Print("Linking in jama Deactivated")
            Downstream_work(values["-IN-"],values["-IN-3"],values["-IN-4"])

################ Menu Bar events   ################################################################   
 
   #### Future implementations ####
    if event == 'New':
        sg.popup_no_buttons('Reserved for future implementations',title='INFO')
    if event == 'options':
 #       sg.popup_no_buttons('Reserved for future implementations',title='INFO')
        sg.Print("client ID=",client_ID)
        sg.Print("client Secret code=",client_code)


    if event == 'fetch':
        #f = open("settings.txt", 'r')
        #settings = json.load(f)
        sg.Print("client ID=",client_ID)
        sg.Print("client Secret code=",client_code)
    if event == 'jama_credentials':
        client_ID=sg.popup_get_text("enter a client ID: ")
        client_code=sg.popup_get_text("enter a client secret code: ")

        sg.Print("client ID=",client_ID)
        sg.Print("client Secret code=",client_code)
        try:
            # Update the JamaClient with new credentials
            jama_client = JamaClient('https://luminar.jamacloud.com', credentials=(client_ID, client_code), oauth=True)
            #data= jama_client.get_items_downstream_related(401046,1)
            #sg.Print(data)
        except:
            sg.popup(title="Error",custom_text="!! Please Enter Correct Credentials !!")
    ###############################

    if event == 'Exit':
        window.close()
    if event == 'Debug Window':
        Debug = 'enable'
        sg.Print('Debugger window started')

    if event == 'About...':
        #print('Current vserion is 1.6')
        sg.popup_no_buttons("Option to add users jama credentials",text_color="Yellow",title="Revision 1_6",non_blocking=True)

    if event == 'Revision History':
        file=open("Rev_His.txt")
        text=file.read()
        sg.popup_scrolled(text, title="Revision History", font=("Arial Bold", 16), size=(50,10))


    if event == 'Contact':
        sg.popup_no_buttons('swapnil.bhujbal@luminartech.com',title="Conatct me @",text_color="orange", non_blocking=True)
##################################################################################################        
window.close()

#sg.Window(title="Jama GUI", layout=[[]], margins=(100, 50)).read()


